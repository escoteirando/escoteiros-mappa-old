from datetime import datetime, timedelta, timezone
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from mappa.models.export.progressao_secao import ProgressaoSecao
from mappa.models.mappa.marcacao import MarcacaoModel
from mappa.models.mappa.progressao import ProgressaoModel
from mappa.models.mappa.secao import SecaoModel
from mappa.models.mappa.subsecao import SubSecaoModel
from mappa.service.mappa_service import MAPPAService


class MAPPAExportService:

    def __init__(self, service: MAPPAService):
        self.svc: MAPPAService = service
        self._progressoes_ramo = {}
        self._marcacoes_secao = {}
        self._secoes = {}
        self._equipe = {}

    def get_secoes(self) -> List[SecaoModel]:
        if not self._secoes:
            self._secoes = self.svc.get_secoes(self.svc._user_id)
        return self._secoes

    def get_progressoes_ramo(self, ramo: str) -> List[ProgressaoModel]:
        if ramo not in self._progressoes_ramo:
            progressoes_ramo = [progressao
                                for progressao in self.svc.get_progressoes()
                                if progressao.ramo == ramo]
            progressoes_ramo.sort(key=lambda progressao: progressao.ordenacao)
            self._progressoes_ramo[ramo] = progressoes_ramo

        return self._progressoes_ramo[ramo]

    def get_marcacoes_secao(self, cod_secao) -> List[MarcacaoModel]:
        if cod_secao not in self._marcacoes_secao:
            marcacoes_secao = self.svc.get_marcacoes(cod_secao)
            self._marcacoes_secao[cod_secao] = marcacoes_secao

        return self._marcacoes_secao[cod_secao]

    def get_equipe(self, cod_secao) -> List[SubSecaoModel]:
        if cod_secao not in self._equipe:
            self._equipe[cod_secao] = self.svc.get_equipe(
                self.svc._user_id, cod_secao)
        return self._equipe[cod_secao]

    def export_progressoes(self, ramo: str, wb: Workbook):
        ws = wb.create_sheet(title='Progressões')
        progressoes = self.get_progressoes_ramo(ramo)
        ws.cell(row=1, column=1, value='Código')
        ws.cell(row=1, column=2, value='Segmento')
        ws.cell(row=1, column=3, value='Descrição')
        progressao: ProgressaoModel
        c1 = c2 = c3 = 0
        for index, progressao in enumerate(progressoes):
            c1 = max(c1, len(str(progressao.codigoUeb)))
            c2 = max(c2, len(progressao.segmento))
            c3 = max(c3, len(progressao.descricao))
            ws.cell(row=index+2, column=1, value=progressao.codigoUeb)
            ws.cell(row=index+2, column=2, value=progressao.segmento)
            ws.cell(row=index+2, column=3, value=progressao.descricao)

        ws.column_dimensions['A'].width = c1
        ws.column_dimensions['B'].width = c2
        ws.column_dimensions['C'].width = c3

    def export_all(self, ramo: str) -> Workbook:
        wb = Workbook()
        ws1 = wb.active
        self.export_progressoes(ramo, wb)
        self.export_secao_progressoes(wb)
        self.export_secao_stats(wb)
        wb.remove(ws1)

        return wb

    def export_secao_progressoes(self, wb: Workbook):
        secoes = self.get_secoes()
        progressoes = {}
        tipo_secao = {1: 'A', 2: 'E', 3: 'S', 4: 'P'}
        limite_idade_secao = {1: 11, 2: 15, 3: 18, 4: 21}
        for secao in secoes:
            progressoes = self.get_progressoes_ramo(
                tipo_secao[secao.codigoTipoSecao])
            marcacoes = self.get_marcacoes_secao(secao.codigo)

            marc_assoc = {}
            for marcacao in marcacoes:
                if marcacao.codigoAssociado not in marc_assoc:
                    marc_assoc[marcacao.codigoAssociado] = set()
                marc_assoc[marcacao.codigoAssociado].add(
                    marcacao.codigoAtividade)

            ws = wb.create_sheet(secao.nome)
            ws.cell(1, 1, 'Subseção')
            ws.cell(1, 3, 'Associado')
            c = ws['A1']
            c.font = Font(bold=True)
            c = ws['C1']
            c.font = Font(bold=True)
            ws.cell(1, 4, 'Progressão %')
            ws.cell(1, 5, 'Nascimento')
            ws.cell(1, 6, 'Tempo restante na seção (m)')

            coluna = 7
            for index, progressao in enumerate(progressoes):
                ws.cell(1, coluna+index, progressao.codigoUeb)
            linha = 2
            equipe = self.get_equipe(secao.codigo)
            for subsecao in equipe:
                for associado in subsecao.associados:
                    ws.cell(row=linha, column=1, value=subsecao.nome)
                    if subsecao.codigoLider == associado.codigo:
                        ws.cell(row=linha, column=2, value='Lider')
                    elif subsecao.codigoViceLider == associado.codigo:
                        ws.cell(row=linha, column=2, value='Vice Lider')
                    ws.cell(row=linha, column=3, value=associado.nome)
                    if associado.codigo in marc_assoc:
                        ws.cell(row=linha, column=4, value=len(
                            marc_assoc[associado.codigo])/len(progressoes))
                    else:
                        ws.cell(row=linha, column=4, value=0)

                    ws.cell(row=linha, column=4).style = 'Percent'

                    ws.cell(row=linha, column=5,
                            value=associado.dataNascimento)
                    ws.cell(row=linha, column=5).number_format = 'dd/mm/yyyy'

                    anos = (datetime.now(timezone.utc) -
                            associado.dataNascimento).days/365.25
                    meses_restantes = int(
                        120 * (limite_idade_secao[secao.codigoTipoSecao] - anos))/10

                    ws.cell(row=linha, column=6, value=meses_restantes)

                    for index, progressao in enumerate(progressoes):
                        if associado.codigo in marc_assoc and\
                                progressao.codigo in marc_assoc[associado.codigo]:
                            ws.cell(row=linha, column=coluna+index, value='X')
                            ws.cell(row=linha, column=coluna +
                                    index).alignment = Alignment(horizontal='center')

                    linha += 1

    def export_secao_stats(self, wb: Workbook):
        secoes = self.get_secoes()

        tipo_secao = {1: 'A', 2: 'E', 3: 'S', 4: 'P'}
        for secao in secoes:
            ws = wb.create_sheet(title="Stats "+secao.nome)
            progressoes = self.get_progressoes_ramo(
                tipo_secao[secao.codigoTipoSecao])
            marcacoes = self.get_marcacoes_secao(secao.codigo)

            stats = {progressao.codigo: ProgressaoSecao(
                0, progressao.codigo, progressao.codigoUeb, progressao.descricao)
                for progressao in progressoes}

            equipe = self.get_equipe(secao.codigo)
            qtd_equipe = 0
            cod_associados = set()
            for subsecao in equipe:
                qtd_equipe += len(subsecao.associados)
                for associado in subsecao.associados:
                    cod_associados.add(associado.codigo)

            associado_por_progressao = {}

            for marcacao in marcacoes:
                if marcacao.codigoAtividade not in associado_por_progressao:
                    associado_por_progressao[marcacao.codigoAtividade] = set()
                if marcacao.codigoAssociado in cod_associados and\
                        marcacao.codigoAssociado not in associado_por_progressao[marcacao.codigoAtividade]:
                    stats[marcacao.codigoAtividade].conquistas += 1
                    associado_por_progressao[marcacao.codigoAtividade].add(
                        marcacao.codigoAssociado)

            ws.cell(row=1, column=1, value='Progressão')
            ws.cell(row=1, column=2, value='Conquistas')
            ws.cell(row=1, column=3, value='Pendências')
            ws.cell(row=1, column=4, value='%')
            ws.cell(row=1, column=5, value='Descrição')

            lista_stats: List[ProgressaoSecao] = list(stats.values())

            # (Código, Conquistas, Descrição)
            lista_stats.sort(key=lambda x: x.conquistas)

            for index, stat in enumerate(lista_stats):
                ws.cell(row=index+2, column=1, value=stat.cod_ueb)
                ws.cell(row=index+2, column=2, value=stat.conquistas)
                ws.cell(row=index+2, column=3,
                        value=qtd_equipe-stat.conquistas)
                ws.cell(row=index+2, column=4,
                        value=stat.conquistas/qtd_equipe)
                ws.cell(row=index+2, column=4).style = "Percent"
                ws.cell(row=index+2, column=5, value=stat.descricao)
